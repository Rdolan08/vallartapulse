#!/usr/bin/env python3
"""
seed-datatur.py
────────────────────────────────────────────────────────────────────────────────
Reads the DATATUR Compendio Estadístico de Turismo en México 2024 (CETM2024.zip)
and replaces the synthetic 2024 rows in tourism_metrics with real figures.

Data extracted from CETM2024/5_1.xlsx:
  • Vista09a  → monthly occupancy rate (fraction, 0–1) for Puerto Vallarta
  • Vista05   → monthly available room-nights for Puerto Vallarta
  • Vista10a  → monthly Average Daily Rate (thousands MXN) for Puerto Vallarta

Run:
  DATABASE_URL=... python3 scripts/seed-datatur.py
  (DATABASE_URL is set automatically in the Replit env)
────────────────────────────────────────────────────────────────────────────────
"""

import os, io, zipfile, sys, math, calendar
import openpyxl
import psycopg2
from datetime import date

DATABASE_URL = os.environ["DATABASE_URL"]
ZIP_PATH     = "/tmp/datatur/CETM2024.zip"
CITY_NAME    = "Puerto Vallarta"
DATA_YEAR    = 2024

# Approximate average MXN/USD for 2024 (Banxico annual average ≈ 17.15)
MXN_USD_2024 = 17.15

MONTHS = [
    "January","February","March","April","May","June",
    "July","August","September","October","November","December"
]

def extract_pvr_row(ws, city=CITY_NAME):
    """Return the 12 monthly values for Puerto Vallarta from a sheet."""
    for row in ws.iter_rows(values_only=True):
        if row[4] and str(row[4]).strip().lower() == city.lower():
            # Columns 7..18 = Jan..Dec (0-indexed: cols index 7 to 18 → positions 7–18)
            # Based on structure: [None, None, None, None, 'City', 'PM?', 'Corridor?', Jan, Feb, ..., Dec, Total, Total]
            vals = [row[i] for i in range(7, 19)]
            return [float(v) if v is not None else None for v in vals]
    return None

def main():
    print(f"Opening {ZIP_PATH} …")
    with zipfile.ZipFile(ZIP_PATH) as z:
        with z.open("CETM2024/5_1.xlsx") as f:
            wb = openpyxl.load_workbook(io.BytesIO(f.read()), read_only=True, data_only=True)
            print("Sheets:", wb.sheetnames)

            # Occupancy rate (Vista09a)
            occ_rates = extract_pvr_row(wb["Vista09a"])
            print(f"Occupancy rates: {[round(r*100, 1) if r else None for r in occ_rates]}")

            # Available room-nights (Vista05)
            avail_rooms = extract_pvr_row(wb["Vista05"])
            print(f"Available room-nights: {avail_rooms}")

            # ADR in thousands MXN (Vista10a)
            adr_k_mxn = extract_pvr_row(wb["Vista10a"])
            print(f"ADR (K MXN): {[round(v, 3) if v else None for v in adr_k_mxn]}")

    if not occ_rates:
        print("ERROR: Puerto Vallarta not found in occupancy sheet", file=sys.stderr)
        sys.exit(1)

    conn = psycopg2.connect(DATABASE_URL)
    cur  = conn.cursor()

    updated = 0
    for i, month_name in enumerate(MONTHS):
        month_num = i + 1
        days      = calendar.monthrange(DATA_YEAR, month_num)[1]

        occ = occ_rates[i]
        avail = avail_rooms[i] if avail_rooms else None
        adr_k = adr_k_mxn[i] if adr_k_mxn else None

        if occ is None:
            print(f"  {month_name}: no occupancy data, skipping")
            continue

        occ_pct = round(occ * 100, 2)

        # Total hotel rooms = available room-nights / days_in_month
        total_rooms = round(avail / days) if avail else None

        # Average daily rate in USD
        avg_rate_usd = round(adr_k * 1000 / MXN_USD_2024, 2) if adr_k else None

        # RevPAR = ADR × Occupancy
        revpar_usd = round(avg_rate_usd * occ, 2) if avg_rate_usd else None

        # Upsert: update existing row or insert if missing
        cur.execute("""
            INSERT INTO tourism_metrics
                (year, month, month_name, hotel_occupancy_rate, total_hotel_rooms,
                 avg_hotel_rate_usd, revenue_per_available_room_usd,
                 international_arrivals, domestic_arrivals, total_arrivals,
                 cruise_visitors, source)
            VALUES (%s,%s,%s,%s,%s,%s,%s,NULL,NULL,NULL,NULL,'DATATUR/SECTUR – Compendio 2024 (real)')
            ON CONFLICT DO NOTHING
        """, (DATA_YEAR, month_num, month_name, occ_pct, total_rooms, avg_rate_usd, revpar_usd))

        rows_inserted = cur.rowcount

        if rows_inserted == 0:
            # Row already exists — update it
            cur.execute("""
                UPDATE tourism_metrics SET
                    hotel_occupancy_rate           = %s,
                    total_hotel_rooms              = %s,
                    avg_hotel_rate_usd             = %s,
                    revenue_per_available_room_usd = %s,
                    source                         = 'DATATUR/SECTUR – Compendio 2024 (real)'
                WHERE year=%s AND month=%s
            """, (occ_pct, total_rooms, avg_rate_usd, revpar_usd, DATA_YEAR, month_num))

        print(f"  {month_name}: occ={occ_pct}%, rooms={total_rooms}, ADR=${avg_rate_usd}, RevPAR=${revpar_usd}")
        updated += 1

    conn.commit()
    cur.close()
    conn.close()
    print(f"\nDone — {updated} months updated in tourism_metrics for {DATA_YEAR}.")

if __name__ == "__main__":
    main()
